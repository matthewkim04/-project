from flask import Flask, request, jsonify
from urlvalidator import validate_url, ValidationError
import openpyxl
import time


app = Flask(__name__)

object_list=["물품목록\n","공구 물품 목록\n","목록\n","물품 목록\n","공구물품목록\n","공구 물품목록\n","물품목록","공구 물품 목록","목록","물품 목록","공구물품목록","공구 물품목록"]
join_list=["참여\n","공구참여\n","공구 참여\n","참여","공구참여","공구 참여"]
enroll_list=["신청\n","공구신청\n","공구 신청\n","물품신청\n","물품 신청\n","신청","공구신청","공구 신청","물품신청","물품 신청"]
name_list=["이름 등록\n","이름 등록","이름등록\n","이름등록","등록\n","등록"]
current_list=["공구 현재 상황","현재 공구 상황","현재","현상황","개인 공구 상황","공구 현재 상황\n","현재 공구 상황\n","현재\n","현상황\n","개인 공구 상황\n"]
delete_enroll_list=["신청 취소","물품 취소","공구 취소","물품 신청 취소","신청 취소\n","물품 취소\n","공구 취소\n","물품 신청 취소\n"]
delete_join_list=["공구 참여 취소","참여 취소","공구참여 취소","공구 참여 취소\n","참여 취소\n","공구참여 취소\n"]
current_list=["현재 상황","개인 상황","개인 현황","현재 상황\n","개인 상황\n","개인 현황\n"]
password="ksa"

ban_list=[]

@app.route('/', methods=['POST'])
def message():
    dataReceive = request.get_json()
    print(dataReceive)
    current_time = time.time()

    if dataReceive['userRequest']['user']['id'] in ban_list:
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "ㅋㅋㄹㅃㅃ 넌 공구하냥에 참여할 수 없닼ㅋㅋ"
                        }
                    }
                ]
            }}

        return jsonify(dataSend)

    if dataReceive["userRequest"]["utterance"] in name_list:
        file=openpyxl.load_workbook("공구하냥.xlsx")
        sheet=file.active
        name=dataReceive["action"]["params"]["text"]

        if dataReceive["action"]["params"]["password"]!=password:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "잘못된 패스워드입니다."
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        i=1
        while True:
            if sheet["O"+str(i)].value==None:
                sheet["O"+str(i)].value=dataReceive['userRequest']['user']['id']
                sheet["P"+str(i)].value=name
                file.save("공구하냥.xlsx")
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이름이 등록되셨습니다.."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O"+str(i)].value==dataReceive['userRequest']['user']['id']:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이미 등록된 이름이 있습니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            i+=1



    if dataReceive["userRequest"]["utterance"] in enroll_list:
        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        url = dataReceive["action"]["params"]["text"]
        party = dataReceive["action"]["params"]["최대 신청 인원수"]
        obj=dataReceive["action"]["params"]["name"]
        id=dataReceive['userRequest']['user']['id']

        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "신청을 하시려면 이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        try:
            x=int(party)
        except ValueError:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "최대 신청 인원수에 숫자만 입력하세요"
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        try:
            validate_url(url)
        except ValidationError:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "유효한 url이 아닙니다!!"
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "신청을 하시려면 이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        i = 1
        while True:
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(int(current_time))
                sheet["B"+str(i)].value=name
                sheet["C" + str(i)].value = url
                sheet["D" + str(i)].value = "1"
                sheet["E" + str(i)].value = party
                sheet["F" + str(i)].value = "T"
                sheet["G"+str(i)].value=obj
                sheet["H"+str(i)].value=name
                file.save("공구하냥.xlsx")
                info = "등록 되었습니다."
                break
            i += 1
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": info
                        }
                    }
                ]
            }}

        return jsonify(dataSend)
    if dataReceive["userRequest"]["utterance"] in join_list:
        flag=True

        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        code=dataReceive["action"]["params"]["text"]
        id = dataReceive['userRequest']['user']['id']
        try:
            x=int(code)
        except ValueError:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "잘못된 고유번호입니다."
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        if int(code)<1 or sheet["A"+str(code)].value==None:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "잘못된 고유번호입니다."
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)
        if int(sheet["A"+str(code)].value)<current_time-172800:
            sheet["F"+str(code)].value = "F"

        if sheet["F" + str(code)].value == "F":
            info = "공구가 종료된 물품입니다."


        elif sheet["F"+str(code)].value == "T":
            if name in sheet["H" + str(code)].value.split('\n'):
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이미 신청하신 물품입니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)

            current_party = str(int(sheet["D" + str(int(code))].value) + 1)
            sheet["H"+str(int(code))].value=sheet["H"+str(int(code))].value+"\n"+name
            sheet["D" + str(int(code))].value = current_party
            if int(sheet["D" + str(int(code))].value) >= int(sheet["E" + str(int(code))].value):
                sheet["F" + str(int(code))].value = "F"
            file.save("공구하냥.xlsx")
            info = "참여 가능합니다. \n관련 공구 문의는 " + sheet["B"+str(int(code))].value + "에게 하면 됩니다."
        else:
            info = "잘못된 고유번호입니다."
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": info
                        }
                    }
                ]
            }}

        return jsonify(dataSend)

    if dataReceive["userRequest"]["utterance"] in object_list:
        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        id = dataReceive['userRequest']['user']['id']
        items = []
        i = 1
        while True:
            if sheet["A" + str(i)].value == None:
                break
            if int(sheet["A"+str(i)].value)<current_time-172800:
                sheet["F"+str(i)].value = "F"
            if int(sheet["D" + str(int(i))].value) >= int(sheet["E" + str(int(i))].value):
                sheet["F" + str(int(i))].value = "F"
            if sheet["F" + str(i)].value == "T" :
                items.append({"title": sheet["G" + str(i)].value, "description": "고유번호: " + str(i) + ",  등록자 수: " +sheet["D" + str(i)].value + "/" +sheet["E" + str(i)].value,"link": {"web": sheet["C" + str(i)].value}})
            i += 1
        if not items:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": '현재 신청된 물품이 없습니다.'
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "listCard": {
                            "header": {
                                "title": "공구 물품 목록"
                            },
                            "items": items
                        }
                    }
                ]
            }
        }

        return jsonify(dataSend)

    if dataReceive["userRequest"]["utterance"] in delete_enroll_list:
        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        obj_code = dataReceive["action"]["params"]["number"]
        id = dataReceive['userRequest']['user']['id']
        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        try:
            x = int(obj_code)
        except ValueError:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "고유번호를 숫자로 입력하세요"
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        if int(obj_code) < 1 or sheet["A" + str(obj_code)].value == None:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "잘못된 고유번호입니다."
                            }
                        }
                    ]
                }}
            return jsonify(dataSend)
        elif sheet["F"+str(obj_code)].value=="F":
            info="이미 취소된 물품입니다."
        elif sheet["B"+str(obj_code)].value!=name:
            info="물품의 신청자가 아니십니다."
        else:
            sheet["F"+str(obj_code)].value="F"
            info="취소가 완료되었습니다."
        file.save("공구하냥.xlsx")
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": info
                        }
                    }
                ]
            }}

        return jsonify(dataSend)

    if dataReceive["userRequest"]["utterance"] in delete_join_list:
        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        obj_code = dataReceive["action"]["params"]["number"]
        id = dataReceive['userRequest']['user']['id']

        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        try:
            x = int(obj_code)
        except ValueError:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "고유번호를 숫자로 입력하세요"
                            }
                        }
                    ]
                }}
            return jsonify(dataSend)

        if int(obj_code) < 1 or sheet["A" + str(obj_code)].value == None:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "잘못된 고유번호입니다."
                            }
                        }
                    ]
                }}
            return jsonify(dataSend)
        elif sheet["F"+str(obj_code)].value=="F":
            info="이미 취소된 물품입니다."
        elif name not in sheet["H"+str(obj_code)].value.split('\n'):
            info="참여하지 않으신 물품입니다."
        elif sheet["B"+str(obj_code)].value==name:
            info="공구 참여 취소가 아니라 물품 신청 취소를 하세요."
        else:
            a=sheet["H"+str(obj_code)].value
            x=a.find(name)
            sheet["H"+str(obj_code)].value=a[:x-1]+a[x+len(name):]
            sheet["D"+str(obj_code)].value=str(int(sheet["D"+str(obj_code)].value)-1)
            info="취소가 완료되었습니다."
        file.save("공구하냥.xlsx")
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": info
                        }
                    }
                ]
            }}

        return jsonify(dataSend)

    if dataReceive["userRequest"]["utterance"] in current_list:
        file = openpyxl.load_workbook("공구하냥.xlsx")
        sheet = file.active
        id = dataReceive['userRequest']['user']['id']

        i=1
        while True:
            if sheet["O" + str(i)].value == None:
                dataSend = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "simpleText": {
                                    "text": "이름을 먼저 등록하셔야 합니다."
                                }
                            }
                        ]
                    }}

                return jsonify(dataSend)
            elif sheet["O" + str(i)].value == id:
                name=sheet["P"+str(i)].value
                break
            i+=1

        i=1
        items=[]
        while True:
            if sheet["A"+str(i)].value==None:
                break
            if int(sheet["A"+str(i)].value)<current_time-172800:
                sheet["F"+str(i)].value = "F"
            if int(sheet["A"+str(i)].value)<current_time-259200:
                i+=1
                continue
            if sheet["B"+str(i)].value==name:
                items.append("물품명: "+sheet["G"+str(i)].value+"\n고유번호: " + str(i) + "\n등록자 수: " +
                             sheet["D" + str(i)].value + "/" + sheet["E" + str(i)].value+"\n링크: "+sheet["C" + str(i)].value+
                             "\n현재 참여한 사람:\n"+sheet["H"+str(i)].value)
                if sheet["F"+str(i)].value=="F":
                    items[-1]="< 종료 >\n"+items[-1]
            i+=1

        if not items:
            dataSend = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "simpleText": {
                                "text": "신청하신 물품이 없습니다."
                            }
                        }
                    ]
                }}

            return jsonify(dataSend)

        stc = ''
        for i in items:
            stc += i + "\n\n"

        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": stc
                        }
                    }
                ]
            }}

        return jsonify(dataSend)

    else:
        dataSend = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "뭔소리야?!!"
                        }
                    }
                ]
            }}

        return jsonify(dataSend)



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
